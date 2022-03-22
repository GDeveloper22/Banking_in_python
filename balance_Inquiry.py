import openpyxl.styles
from colorama import Fore, Style, init

init(autoreset=True)
ws = openpyxl.load_workbook("./Account Database.xlsx")
wsa = ws.active


def bal_Inq():
    try:
        global ac_row
        while True:
            print(Style.DIM + Fore.WHITE + "Enter your account number: ", end="" + Style.RESET_ALL)
            acc_number = input()
            if len(acc_number) == 7:
                if acc_number.isnumeric():
                    for i in range(2, wsa.max_row+1):
                        if acc_number == str(wsa.cell(i, 1).value):
                            ac_row = i
                            max_attempt = 3
                            while max_attempt > 0:
                                print(Style.DIM + Fore.WHITE + "Enter the access pin: ", end="" + Style.RESET_ALL)
                                acc_pin_ = input()
                                if acc_pin_.isnumeric():
                                    if len(acc_pin_) == 4:
                                        if acc_pin_ == str(wsa.cell(ac_row, 2).value):
                                            bal = str(wsa.cell(ac_row, 14).value)
                                            print(Style.DIM + Fore.GREEN + f"Currently you have ${bal} in your account.")
                                            return 0
                                        else:
                                            max_attempt -= 1
                                            if max_attempt == 0:
                                                print(Style.DIM + Fore.RED + f"({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                            else:
                                                print(Style.DIM + Fore.RED + f"Please enter correct access pin ({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                    else:
                                        print(Style.DIM + Fore.RED + "The pin should be 4 digit only..." + Style.RESET_ALL)
                                else:
                                    print(Style.DIM + Fore.RED + "The pin should contains only digit..." + Style.RESET_ALL)
                            print(Style.DIM + Fore.RED + "Out of attempts..." + Style.RESET_ALL)
                            return 0
                    print(Style.DIM + Fore.RED + "Account number not found..." + Style.RESET_ALL)
                else:
                    print(Style.DIM + Fore.RED + "Account number must contain digits only...." + Style.RESET_ALL)
            else:
                print(Style.DIM + Fore.RED + "Account number must be of 7 digits...." + Style.RESET_ALL)
    except Exception as e:
        print("Error found!!, ", e)
