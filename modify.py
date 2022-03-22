import datetime
import openpyxl.styles
from colorama import Fore, Style, init

init(autoreset=True)
ws = openpyxl.load_workbook("./Account Database.xlsx")
wsa = ws.active
wk = openpyxl.load_workbook("./Country-codes.xlsx")
wka = wk.active

global f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no, T_C, acc_row, ac_row


def modify_Acc():
    try:
        global ac_row
        while True:
            print(Style.DIM + Fore.WHITE + "Enter your account number: ", end="" + Style.RESET_ALL)
            acc_number = input()
            if len(acc_number) == 7:
                if acc_number.isnumeric():
                    for i in range(2, wsa.max_row + 1):
                        if acc_number == str(wsa.cell(i, 1).value):
                            ac_row = i
                            max_attempt = 3
                            while max_attempt > 0:
                                print(Style.DIM + Fore.WHITE + "Enter the access pin: ", end="" + Style.RESET_ALL)
                                acc_pin_ = input()
                                if acc_pin_.isnumeric():
                                    if len(acc_pin_) == 4:
                                        if acc_pin_ == str(wsa.cell(ac_row, 2).value):
                                            while True:
                                                print(Style.BRIGHT + Fore.BLUE + "Please choose which thing you want to change" + Style.BRIGHT + Fore.MAGENTA + "\nHere is the menu containing the different options:")
                                                print(Style.BRIGHT + Fore.MAGENTA + "\t1) Change First Name\n\t2) Change Middle name\n\t3) Change Last Name\n\t4) Change Gender\n\t5) Change Email\n\t6) Change Phone No\n\t7) Change Date of Birth\n\t8)User Access Pin\n\t9)Employment Status\n\t10)Account Status")
                                                print(Style.BRIGHT + Fore.GREEN + "Please choose any one option from above: " + Style.RESET_ALL, end="")
                                                q = int(input())
                                                if q == 1:
                                                    first_name()
                                                    break
                                                elif q == 2:
                                                    midd_name()
                                                    break
                                                elif q == 3:
                                                    last_name()
                                                    break
                                                elif q == 4:
                                                    gend()
                                                    break
                                                elif q == 5:
                                                    mail()
                                                    break
                                                elif q == 6:
                                                    phone()
                                                    break
                                                elif q == 7:
                                                    d_o_b()
                                                    age_calc(dob)
                                                    break
                                                elif q == 8:
                                                    pin()
                                                    break
                                                elif q == 9:
                                                    employment()
                                                    break
                                                elif q == 10:
                                                    Acc_stat()
                                                    break
                                                else:
                                                    print(Style.DIM + Fore.RED + "Please choose valid option\n")
                                            return verify()
                                        else:
                                            max_attempt -= 1
                                            if max_attempt == 0:
                                                print(Style.DIM + Fore.RED + f"({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                            else:
                                                print(Style.DIM + Fore.RED + f"Please enter correct access pin ({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                    else:
                                        print(Style.DIM + Fore.RED + "The pin should be 4 digit only..." + Style.RESET_ALL)
                                else:
                                    print(Style.DIM + Fore.RED + "The pin should contains only digit..." + Style.RESET_ALL)
                            print(Style.DIM + Fore.RED + "Out of attempts..." + Style.RESET_ALL)
                            return 0
                    print(Style.DIM + Fore.RED + "Account number not found..." + Style.RESET_ALL)
                else:
                    print(Style.DIM + Fore.RED + "Account number must contain digits only...." + Style.RESET_ALL)
            else:
                print(Style.DIM + Fore.RED + "Account number must be of 7 digits...." + Style.RESET_ALL)
    except Exception as e:
        print("Error found!!, Error: ", e)


def verify():
    try:
        while True:
            print(Style.BRIGHT + Fore.GREEN + "Please choose (Y) or (N) if the above information are  correct or not: ", end="")
            Y_N = input()
            if Y_N.upper() == "Y":
                save()
                return Y_N  # move to Terms and conditions
            elif Y_N.upper() == "N":
                modify_Acc()  # ask which details of are false by showing the menu
                return Y_N
            else:
                print(Style.BRIGHT + Fore.RED + "Please enter valid input...")
    except Exception as e:
        print("Error!!!, ", e)


def save():
    try:
        mx_row = wsa.max_row
        cur_row = ac_row
        try:
            wsa.cell(cur_row, 2).value = acc_pin  # user access pin
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 3).value = acc_status  # account status
            if acc_status == "Active":
                wsa.cell(cur_row, 3).fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
            if acc_status == "Inactive":
                wsa.cell(cur_row, 3).fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 4).value = f_name
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 5).value = m_name
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 6).value = l_name
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 7).value = age
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 8).value = gender
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 9).value = email
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 10).value = pn
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 11).value = dob
        except Exception:
            pass

        try:
            wsa.cell(cur_row, 15).value = emp_status
        except Exception:
            pass

        ws.save("Account Database.xlsx")
        print(Style.BRIGHT + Fore.CYAN + "Your account have been Modified.")
    except Exception as e:
        print("Error!!, ", e)


def first_name():
    global f_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your first name: ", end="")
            f_name = input()
            if f_name.isalpha():
                return f_name
            else:
                if f_name == "":
                    print(Style.DIM + Fore.RED + "First Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        first_name()


def midd_name():
    global m_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your middle name: ", end="")
            m_name = input()
            if m_name.isalpha():
                return m_name
            elif m_name == "":
                m_name = "None"
                return m_name
            else:
                print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        midd_name()


def last_name():
    global l_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your last name: ", end="")
            l_name = input()
            if l_name.isalpha():
                return l_name
            else:
                if l_name == "":
                    print(Style.DIM + Fore.RED + "Last Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        last_name()


def d_o_b():
    global dob
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter date of birth in (YYYY-MM-DD) format: ", end="")
            dob = input()
            dt = datetime.date.today()
            if (int(dob[0:4]) <= int(dt.year)) and (len(dob) == 10):
                if (dob[4] == "-") and (dob[7] == "-"):
                    if int(dob[5:7]) <= 12:
                        if (int(dob[8:10]) <= 30) or (int(dob[8:10]) <= 31):
                            return dob
                        else:
                            print(Style.DIM + Fore.RED + "Please enter valid Birth Date...")
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid Birth Month...")
                else:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
            else:
                if len(dob) != 10:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
                else:
                    print(Style.DIM + Fore.RED + "Please enter valid Birth Year...")
    except Exception as e:
        print("Error!!!, ", e)
        d_o_b()


def age_calc(dob):
    global age
    try:
        dt = datetime.date.today()
        age = int(dt.year) - int(dob[0:4]) - (((int(dt.month)), int(dt.day)) < ((int(dob[5:7])), int(dob[8:10])))  # Subtract today's year to birth year, then compare current month & birthdate to birth month and birthdate
        return age
    except Exception as e:
        print("Error!!!, ", e)
        age_calc(dob)


def gend():
    global gender
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your gender (M)-Male, (F)-Female, (U)-Unspecified: ", end="")
            gender = input()
            if gender.upper() in ("M", "F", "U"):
                return gender
            else:
                print(Style.DIM + Fore.RED + "Please specify proper gender...")
    except Exception as e:
        print("Error!!!, ", e)
        gend()


def mail():
    global email
    try:
        email_suffix = ["gmail.com", "yahoo.com", "hotmail.com", "aol.com", "hotmail.co.uk", "hotmail.fr", "msn.com",
                        "yahoo.fr", "wanadoo.fr", "orange.fr", "comcast.net", "yahoo.co.uk", "yahoo.com.br",
                        "yahoo.co.in", "live.com", "rediffmail.com",
                        "free.fr", "gmx.de", "web.de", "yandex.ru", "ymail.com", "libero.it", "outlook.com",
                        "outlook.ca", "uol.com.br", "bol.com.br", "mail.ru", "cox.net", "hotmail.it", "sbcglobal.net",
                        "sfr.fr", "live.fr", "verizon.net", "live.co.uk", "googlemail.com",
                        "yahoo.es", "ig.com.br", "live.nl", "bigpond.com", "terra.com.br", "yahoo.it", "neuf.fr",
                        "yahoo.de", "alice.it", "rocketmail.com", "att.net", "laposte.net", "facebook.com",
                        "bellsouth.net", "yahoo.in", "hotmail.es", "charter.net",
                        "yahoo.ca", "yahoo.com.au", "rambler.ru", "hotmail.de", "tiscali.it", "shaw.ca", "yahoo.co.jp",
                        "sky.com", "earthlink.net", "optonline.net", "freenet.de", "t-online.de", "aliceadsl.fr",
                        "virgilio.it", "home.nl", "qq.com", "telenet.be",
                        "me.com", "yahoo.com.ar", "tiscali.co.uk", "yahoo.com.mx", "voila.fr", "gmx.net", "mail.com",
                        "planet.nl", "tin.it", "live.it", "ntlworld.com", "arcor.de", "yahoo.co.id", "frontiernet.net",
                        "hetnet.nl", "live.com.au", "yahoo.com.sg",
                        "zonnet.nl", "club-internet.fr", "juno.com", "optusnet.com.au", "blueyonder.co.uk",
                        "bluewin.ch", "skynet.be", "sympatico.ca", "windstream.net", "mac.com", "centurytel.net",
                        "chello.nl", "live.ca", "aim.com", "bigpond.net.au"]
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your email: ", end="")
            email = input()
            if "@" in email:
                if email.split("@", 2)[1] in email_suffix:
                    sp_c = ["!", "@", "#", "$", "%", "^", "&", "*", "(", ")", "_", "+"]
                    while True:
                        if not (any(em in email.split("@", 2)[0] for em in sp_c)):
                            return email
                        else:
                            print(Style.DIM + Fore.RED + "Email Suffix can't contain special characters...")
                else:
                    print(Style.DIM + Fore.RED + "Please specify email with proper prefix and suffix...")
            else:
                if "@" not in email:
                    print(Style.DIM + Fore.RED + "Email contains @ as suffix...")
    except Exception as e:
        print("Error!!!, ", e)
        mail()


def phone():
    global pn
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter country code with (+): ", end="")
            p_cd = input()
            cde = {}
            i = 1
            while i < (wka.max_row - 1):
                cde[wka.cell(i, 1).value] = wka.cell(i, 3).value
                i += 1
            if p_cd in cde.values():
                while True:
                    print(Style.BRIGHT + Fore.WHITE + "Enter your 10 digit phone number: ", end="")
                    p_num = input()
                    if len(p_num) == 10:
                        if p_num.isnumeric():
                            pn = str(p_cd) + str(p_num)
                            return pn
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid phone number...")
            else:
                print(Style.DIM + Fore.RED + "Please enter valid country code...")
    except Exception as e:
        print("Error!!!, ", e)
        phone()


def employment():
    global emp_status
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your Employment Status (Student, Self-Employed, Job, UnEmployed): ", end="")
            emp_status = input()
            if emp_status.capitalize() in ("Student", "Self-employed", "Job", "Unemployed"):
                return emp_status
            else:
                print(Style.DIM + Fore.RED + "Please enter valid Employment Status...")
    except Exception as e:
        print("Error!!!, ", e)
        employment()


def pin():
    global acc_pin
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter 4 digit access pin for security and banking purpose: ", end="")
            acc_pin = input()
            if acc_pin.isnumeric():
                if len(acc_pin) == 4:
                    if acc_pin[0] != acc_pin[1] != acc_pin[2] != acc_pin[3]:
                        return acc_pin
                    else:
                        print(Style.DIM + Fore.RED + "The pin should be odd not in the sequential order like 0000,1111...")
                else:
                    print(Style.DIM + Fore.RED + "The pin should be 4 digit only...")
            else:
                print(Style.DIM + Fore.RED + "The pin should contains only digit...")
    except Exception as e:
        print("Error!!!, ", e)
        pin()


def Acc_stat():
    global acc_status
    while True:
        print(Style.BRIGHT + Fore.WHITE + "Choose Account status 1). Active or 2).Inactive: ", end="")
        stat = input()
        if stat.upper() == '1':
            acc_status = "Active"
            break
        elif stat.upper() == '2':
            acc_status = "Inactive"
            break
        else:
            print(Style.DIM + Fore.RED + "Please enter valid status..." + Style.RESET_ALL)
