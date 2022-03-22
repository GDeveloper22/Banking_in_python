import datetime
import openpyxl.styles
from colorama import Fore, Style, init
import random


init(autoreset=True)
ws = openpyxl.load_workbook("./Account Database.xlsx")
wsa = ws.active
wk = openpyxl.load_workbook("./Country-codes.xlsx")
wka = wk.active

global f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no, T_C, acc_row, ac_row


def welcome():
    print(Style.BRIGHT + Fore.YELLOW + "\nWelcome to the Swiss Bank")
    print(Style.BRIGHT + Fore.RED + "To open a new account please follow the steps given below")
    print(Style.BRIGHT + Fore.RED + "NOTE: For the Opening an account & identity verification, we would like to ask your personal details")


def first_name():
    global f_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your first name: ", end="")
            f_name = input()
            if f_name.isalpha():
                return f_name
            else:
                if f_name == "":
                    print(Style.DIM + Fore.RED + "First Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        first_name()


def midd_name():
    global m_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your middle name: ", end="")
            m_name = input()
            if m_name.isalpha():
                return m_name
            elif m_name == "":
                m_name = "None"
                return m_name
            else:
                print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        midd_name()


def last_name():
    global l_name
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Please enter your last name: ", end="")
            l_name = input()
            if l_name.isalpha():
                return l_name
            else:
                if l_name == "":
                    print(Style.DIM + Fore.RED + "Last Name can not be empty")
                else:
                    print(Style.DIM + Fore.RED + "Name only contains letters...")
    except Exception as e:
        print("Error!!!, ", e)
        last_name()


def d_o_b():
    global dob
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter date of birth in (YYYY-MM-DD) format: ", end="")
            dob = input()
            dt = datetime.date.today()
            if (int(dob[0:4]) <= int(dt.year)) and (len(dob) == 10):
                if (dob[4] == "-") and (dob[7] == "-"):
                    if int(dob[5:7]) <= 12:
                        if (int(dob[8:10]) <= 30) or (int(dob[8:10]) <= 31):
                            return dob
                        else:
                            print(Style.DIM + Fore.RED + "Please enter valid Birth Date...")
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid Birth Month...")
                else:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
            else:
                if len(dob) != 10:
                    print(Style.DIM + Fore.RED + "Enter date of birth with valid format")
                else:
                    print(Style.DIM + Fore.RED + "Please enter valid Birth Year...")
    except Exception as e:
        print("Error!!!, ", e)
        d_o_b()


def age_calc(dob):
    global age
    try:
        dt = datetime.date.today()
        age = int(dt.year) - int(dob[0:4]) - (((int(dt.month)), int(dt.day)) < ((int(dob[5:7])), int(dob[8:10])))  # Subtract today's year to birth year, then compare current month & birthdate to birth month and birthdate
        return age
    except Exception as e:
        print("Error!!!, ", e)
        age_calc(dob)


def gend():
    global gender
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your gender (M)-Male, (F)-Female, (U)-Unspecified: ", end="")
            gender = input()
            if gender.upper() in ("M", "F", "U"):
                return gender
            else:
                print(Style.DIM + Fore.RED + "Please specify proper gender...")
    except Exception as e:
        print("Error!!!, ", e)
        gend()


def mail():
    global email
    try:
        email_suffix = ["gmail.com", "yahoo.com", "hotmail.com", "aol.com", "hotmail.co.uk", "hotmail.fr", "msn.com",
                        "yahoo.fr", "wanadoo.fr", "orange.fr", "comcast.net", "yahoo.co.uk", "yahoo.com.br",
                        "yahoo.co.in", "live.com", "rediffmail.com",
                        "free.fr", "gmx.de", "web.de", "yandex.ru", "ymail.com", "libero.it", "outlook.com",
                        "outlook.ca", "uol.com.br", "bol.com.br", "mail.ru", "cox.net", "hotmail.it", "sbcglobal.net",
                        "sfr.fr", "live.fr", "verizon.net", "live.co.uk", "googlemail.com",
                        "yahoo.es", "ig.com.br", "live.nl", "bigpond.com", "terra.com.br", "yahoo.it", "neuf.fr",
                        "yahoo.de", "alice.it", "rocketmail.com", "att.net", "laposte.net", "facebook.com",
                        "bellsouth.net", "yahoo.in", "hotmail.es", "charter.net",
                        "yahoo.ca", "yahoo.com.au", "rambler.ru", "hotmail.de", "tiscali.it", "shaw.ca", "yahoo.co.jp",
                        "sky.com", "earthlink.net", "optonline.net", "freenet.de", "t-online.de", "aliceadsl.fr",
                        "virgilio.it", "home.nl", "qq.com", "telenet.be",
                        "me.com", "yahoo.com.ar", "tiscali.co.uk", "yahoo.com.mx", "voila.fr", "gmx.net", "mail.com",
                        "planet.nl", "tin.it", "live.it", "ntlworld.com", "arcor.de", "yahoo.co.id", "frontiernet.net",
                        "hetnet.nl", "live.com.au", "yahoo.com.sg",
                        "zonnet.nl", "club-internet.fr", "juno.com", "optusnet.com.au", "blueyonder.co.uk",
                        "bluewin.ch", "skynet.be", "sympatico.ca", "windstream.net", "mac.com", "centurytel.net",
                        "chello.nl", "live.ca", "aim.com", "bigpond.net.au"]
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your email: ", end="")
            email = input()
            if "@" in email:
                if email.split("@", 2)[1] in email_suffix:
                    sp_c = ["!", "@", "#", "$", "%", "^", "&", "*", "(", ")", "_", "+"]
                    while True:
                        if not (any(em in email.split("@", 2)[0] for em in sp_c)):
                            return email
                        else:
                            print(Style.DIM + Fore.RED + "Email Suffix can't contain special characters...")
                else:
                    print(Style.DIM + Fore.RED + "Please specify email with proper prefix and suffix...")
            else:
                if "@" not in email:
                    print(Style.DIM + Fore.RED + "Email contains @ as suffix...")
    except Exception as e:
        print("Error!!!, ", e)
        mail()


def phone():
    global pn
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter country code with (+): ", end="")
            p_cd = input()
            cde = {}
            i = 1
            while i < (wka.max_row - 1):
                cde[wka.cell(i, 1).value] = wka.cell(i, 3).value
                i += 1
            if p_cd in cde.values():
                while True:
                    print(Style.BRIGHT + Fore.WHITE + "Enter your 10 digit phone number: ", end="")
                    p_num = input()
                    if len(p_num) == 10:
                        if p_num.isnumeric():
                            pn = str(p_cd) + str(p_num)
                            return pn
                    else:
                        print(Style.DIM + Fore.RED + "Please enter valid phone number...")
            else:
                print(Style.DIM + Fore.RED + "Please enter valid country code...")
    except Exception as e:
        print("Error!!!, ", e)
        phone()


def employment():
    global emp_status
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter your Employment Status (Student, Self-Employed, Job, UnEmployed): ", end="")
            emp_status = input()
            if emp_status.capitalize() in ("Student", "Self-employed", "Job", "Unemployed"):
                return emp_status
            else:
                print(Style.DIM + Fore.RED + "Please enter valid Employment Status...")
    except Exception as e:
        print("Error!!!, ", e)
        employment()


def toc():
    global acc_type
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter the type of account that you want to open \n(C) for checking (S) for savings, (MM) for Money Market Account: ", end="")
            acc_type = input()
            if acc_type.upper() in ("C", "S", "MM"):
                return acc_type
            else:
                print(Style.DIM + Fore.RED + "Please choose valid account type from above...")
    except Exception as e:
        print("Error!!!, ", e)
        toc()


def init_dep():
    global ini_dep
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter the initial deposit amount for account opening: $", end="")
            ini_dep = input()
            if ini_dep.isnumeric():
                if int(ini_dep) >= 10:
                    return ini_dep
                else:
                    print(Style.DIM + Fore.RED + "The minimum amount to deposit is $10...")
    except Exception as e:
        print("Error!!!, ", e)
        init_dep()


def pin():
    global acc_pin
    try:
        while True:
            print(Style.BRIGHT + Fore.WHITE + "Enter 4 digit access pin for security and banking purpose: ", end="")
            acc_pin = input()
            if acc_pin.isnumeric():
                if len(acc_pin) == 4:
                    if acc_pin[0] != acc_pin[1] != acc_pin[2] != acc_pin[3]:
                        return acc_pin
                    else:
                        print(Style.DIM + Fore.RED + "The pin should be odd not in the sequential order like 0000,1111...")
                else:
                    print(Style.DIM + Fore.RED + "The pin should be 4 digit only...")
            else:
                print(Style.DIM + Fore.RED + "The pin should contains only digit...")
    except Exception as e:
        print("Error!!!, ", e)
        pin()


def Acc_stat():
    global acc_status
    acc_status = "Active"
    return acc_status


def gen_acc():
    global acc_no
    try:
        while True:
            acc_no = random.randint(1111111, 9999999)
            i = 2
            acc_list = []
            while i < wsa.max_row:
                acc_no_database = wsa.cell(i, 1)
                acc_list.append(acc_no_database)
                i += 1
            if acc_no not in acc_list:
                return acc_no
    except Exception as e:
        print("Error!!!, ", e)
        gen_acc()


def details(f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no):
    try:
        print(Style.BRIGHT + Fore.BLUE + "Please verify all of your details:")
        print(Style.BRIGHT + Fore.BLUE + "First Name: " + Style.DIM + Fore.YELLOW + f_name)
        print(Style.BRIGHT + Fore.BLUE + "Middle Name: " + Style.DIM + Fore.YELLOW + m_name)
        print(Style.BRIGHT + Fore.BLUE + "Last Name: " + Style.DIM + Fore.YELLOW + l_name)
        print(Style.BRIGHT + Fore.BLUE + "Date of Birth: " + Style.DIM + Fore.YELLOW + dob)
        print(Style.BRIGHT + Fore.BLUE + "Age: " + Style.DIM + Fore.YELLOW + str(age))
        print(Style.BRIGHT + Fore.BLUE + "Gender: " + Style.DIM + Fore.YELLOW + gender)
        print(Style.BRIGHT + Fore.BLUE + "Email: " + Style.DIM + Fore.YELLOW + email)
        print(Style.BRIGHT + Fore.BLUE + "Phone number: " + Style.DIM + Fore.YELLOW + pn)
        print(Style.BRIGHT + Fore.BLUE + "Employment status: " + Style.DIM + Fore.YELLOW + emp_status)
        print(Style.BRIGHT + Fore.BLUE + "Type of Account: " + Style.DIM + Fore.YELLOW + acc_type)
        print(Style.BRIGHT + Fore.BLUE + "Initial deposit while opening an account: " + Style.DIM + Fore.YELLOW + ini_dep)
        print(Style.BRIGHT + Fore.BLUE + "User access pin: " + Style.DIM + Fore.YELLOW + acc_pin)
        print(Style.BRIGHT + Fore.BLUE + "Account Status: " + Style.DIM + Fore.YELLOW + acc_status)
        print(Style.BRIGHT + Fore.BLUE + "Account Number: " + Style.DIM + Fore.YELLOW + str(acc_no))
        print(Style.DIM + Fore.RED + "Please ensure that all the details that you have provided must verify with your identity\n")
        return verify()
    except Exception as e:
        print("Error!!!, ", e)
        details(f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no)


# Ask if the details are correct, y== go to next step, no=> ask to choose which step is wrong
def verify():
    try:
        while True:
            print(Style.BRIGHT + Fore.GREEN + "Please choose (Y) or (N) if the above information are  correct or not: ", end="")
            Y_N = input()
            if Y_N.upper() == "Y":
                t_c()
                return Y_N  # move to Terms and conditions
            elif Y_N.upper() == "N":
                new_Acc_change()  # ask which details of are false by showing the menu
                return Y_N
            else:
                print(Style.BRIGHT + Fore.RED + "Please enter valid input...")
    except Exception as e:
        print("Error!!!, ", e)
        verify()


# Print statement which ask user to accept the Terms and Conditions for opening an account with swiss bank. If y== store the details in the Excel and print message creating account... and after that display message which shows that we are reviewing your account, and you will get updated in 2-3 business days. If no== Print message, sorry you are ineligible to open an account with us as you are not agreeing with our T&C
def t_c():
    try:
        while True:
            global T_C
            print(Style.DIM + Fore.RED + "\nPlease press (Y) to agree with our " + Style.BRIGHT + Fore.GREEN + "TERMS AND CONDITION" + Style.DIM + Fore.RED + " OR enter (N) to deny " + Style.RESET_ALL)
            print(Style.DIM + Fore.RED + "NOTE: By pressing (N) your account will not be created and all of the information you have entered above will be erased " + Style.RESET_ALL, end="")
            T_C = input()
            if T_C.upper() == "Y":
                saving()  # proceed further
            elif T_C.upper() == "N":
                print(Style.BRIGHT + Fore.GREEN + "Are you sure that you are not agreeing with our T&C's: " + Style.RESET_ALL, end="")
                D_V = input()
                if D_V.upper() == "Y":
                    print("Sorry! As you are not agreeing with our terms and conditions we can not proceed further.")  # ask which details of are false by showing the menu
                    return T_C  # cancel account creating process
                elif D_V.upper() == "N":
                    print("Account creation under progress")
                    T_C = "Y"
                    saving()  # continue the account creation process
                else:
                    print(Style.DIM + Fore.RED + "Please enter valid input...")
                break
            else:
                print(Style.DIM + Fore.RED + "Please enter valid input...")
    except Exception as e:
        print("Error!!!, ", e)
        t_c()


def new_Acc():
    try:
        welcome()
    # First name
        first_name()
    # Middle Name
        midd_name()
    # Last Name
        last_name()
    # DOB
        d_o_b()
    # Age (will be calculated according to DOB)
        age_calc(dob)
    # Gender
        gend()
    # Email
        mail()
    # Phone Number
        phone()
    # Employment status (Self-Employed, Job, Unemployed, Student)
        employment()
    # type of account (C) for checking (S) for savings, (MM) for Money Market Account
        toc()
    # initial deposit (Min $10)
        init_dep()
    # Set user access pin for NET Banking (4 digit, it should not be sequence like 0000, 1111)
        pin()
    # Account Status (Active by default while opening)
        Acc_stat()
        # color the cell of the Excel with green color if active or red when inactive and also set font color as white.
    # Now generate a random account number of 7 digit & when there is a new user then it will not generate the same number, as of other one.
        gen_acc()
    # Lastly show all details
        details(f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no)
    except Exception as e:
        print("Error found!!, Error: ", e)


def saving():
    try:
        mx_row = wsa.max_row
        cur_row = mx_row + 1
        wsa.cell(cur_row, 1).value = acc_no  # Account no
        wsa.cell(cur_row, 2).value = acc_pin  # user access pin
        wsa.cell(cur_row, 3).value = acc_status  # account status
        if acc_status == "Active":
            wsa.cell(cur_row, 3).fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
        if acc_status == "Inactive":
            wsa.cell(cur_row, 3).fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
        wsa.cell(cur_row, 4).value = f_name
        wsa.cell(cur_row, 5).value = m_name
        wsa.cell(cur_row, 6).value = l_name
        wsa.cell(cur_row, 7).value = age
        wsa.cell(cur_row, 8).value = gender
        wsa.cell(cur_row, 9).value = email
        wsa.cell(cur_row, 10).value = pn
        wsa.cell(cur_row, 11).value = dob
        wsa.cell(cur_row, 12).value = acc_type
        wsa.cell(cur_row, 13).value = ini_dep
        wsa.cell(cur_row, 14).value = ini_dep
        wsa.cell(cur_row, 15).value = emp_status
        wsa.cell(cur_row, 16).value = f"Your Account has been created and you have deposited {ini_dep}."  # Account Activity
        wsa.cell(cur_row, 17).value = f"{datetime.date.today()}"  # Add account opening date and time.
        wsa.cell(cur_row, 18).value = T_C
        ws.save("Account Database.xlsx")
        print(Style.BRIGHT + Fore.CYAN + "Hurray!, Your account have been created.")
        return 0
    except Exception as e:
        print("Error!!, ", e)


def new_Acc_change():
    try:
        while True:
            print(Style.BRIGHT + Fore.BLUE + "Please choose which thing you want to change" + Style.BRIGHT + Fore.MAGENTA + "\nHere is the menu containing the different options:")
            print(Style.BRIGHT + Fore.MAGENTA + "\t1) Change First Name\n\t2) Change Middle name\n\t3) Change Last Name\n\t4) Change Gender\n\t5) Change Email\n\t6) Change Phone No\n\t7) Change Date of Birth\n\t8) Type of Account\n\t9) Initial deposit\n\t10)User Access Pin\n\t11).Employment Status")
            print(Style.BRIGHT + Fore.GREEN + "Please choose any one option from above: ")
            print(Style.BRIGHT + Fore.WHITE + "Enter number for changing its corresponding value: ", end="")
            q = int(input())
            if q == 1:
                first_name()
                break
            elif q == 2:
                midd_name()
                break
            elif q == 3:
                last_name()
                break
            elif q == 4:
                gend()
                break
            elif q == 5:
                mail()
                break
            elif q == 6:
                phone()
                break
            elif q == 7:
                d_o_b()
                age_calc(dob)
                break
            elif q == 8:
                toc()
                break
            elif q == 9:
                init_dep()
                break
            elif q == 10:
                pin()
                break
            elif q == 11:
                employment()
                break
            else:
                print(Style.DIM + Fore.RED + "Please choose valid option\n")
        return details(f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no)
    except Exception as e:
        print("Error found!!, ", e)
        new_Acc_change()  # enter a function which will ask to enter again to input the data.
