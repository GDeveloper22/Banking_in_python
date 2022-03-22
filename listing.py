import openpyxl.styles
from colorama import Fore, Style, init
import time

init(autoreset=True)
ws = openpyxl.load_workbook("./Account Database.xlsx")
wsa = ws.active
wk = openpyxl.load_workbook("./Country-codes.xlsx")
wka = wk.active

Authorization_list = ["6249263"]  # list of account number which has access to all the data like manager and owner.
bank_authorization = 10032022

global f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no, T_C, acc_row, ac_row


def list_Acc():  # Can be accessed only by manager or bank security
    try:
        global ac_row
        while True:
            print(Style.DIM + Fore.WHITE + "Enter your account number: ", end="" + Style.RESET_ALL)
            acc_number = input()
            if len(acc_number) == 7:
                if acc_number.isnumeric():
                    for i in range(2, wsa.max_row+1):
                        if acc_number == str(wsa.cell(i, 1).value):
                            ac_row = i
                            max_attempt = 3
                            if acc_number in Authorization_list:
                                while max_attempt > 0:
                                    print(Style.DIM + Fore.WHITE + "Enter the access pin: ", end="" + Style.RESET_ALL)
                                    acc_pin_ = input()
                                    if acc_pin_.isnumeric():
                                        if len(acc_pin_) == 4:
                                            if acc_pin_ == str(wsa.cell(ac_row, 2).value):
                                                att = 3
                                                while att > 0:
                                                    print(Style.DIM + Fore.WHITE + "Enter the Bank authorization code: ", end="" + Style.RESET_ALL)
                                                    auth_code = input()
                                                    if auth_code.isnumeric():
                                                        if len(auth_code) == 8:
                                                            if int(auth_code) == bank_authorization:
                                                                # run loop, and save details in the variable and then print again run loop
                                                                x = 1  # row
                                                                while x <= wsa.max_row:
                                                                    y = 1  # column
                                                                    while y < 16:
                                                                        print(Style.BRIGHT + Fore.CYAN + str(wsa.cell(x, y).value).ljust(22), end="" + Style.RESET_ALL)
                                                                        y += 1
                                                                    x += 1
                                                                    print()
                                                                return 0
                                                            else:
                                                                att -= 1
                                                                if att == 0:
                                                                    print(Style.DIM + Fore.RED + f"({att} Attempts left)..." + Style.RESET_ALL)
                                                                    return 0
                                                                else:
                                                                    print(Style.DIM + Fore.RED + f"Please enter correct access code ({att} Attempts left)..." + Style.RESET_ALL)
                                                        else:
                                                            print(Style.DIM + Fore.RED + "Code must be of 8 digits...." + Style.RESET_ALL)
                                                    else:
                                                        print(Style.DIM + Fore.RED + "Code must contain digits only...." + Style.RESET_ALL)
                                            else:
                                                max_attempt -= 1
                                                if max_attempt == 0:
                                                    print(Style.DIM + Fore.RED + f"({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                                    return 0
                                                else:
                                                    print(Style.DIM + Fore.RED + f"Please enter correct access pin ({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                        else:
                                            print(Style.DIM + Fore.RED + "The pin should be 4 digit only..." + Style.RESET_ALL)
                                    else:
                                        print(Style.DIM + Fore.RED + "The pin should contains only digit..." + Style.RESET_ALL)
                                print(Style.DIM + Fore.RED + "Out of attempts..." + Style.RESET_ALL)
                                return 0
                            else:
                                print(Style.DIM + Fore.RED + "You are not authorized for accessing this function..." + Style.RESET_ALL)
                                return 0
                    print(Style.DIM + Fore.RED + "Account number not found..." + Style.RESET_ALL)
                else:
                    print(Style.DIM + Fore.RED + "Account number must contain digits only...." + Style.RESET_ALL)
            else:
                print(Style.DIM + Fore.RED + "Account number must be of 7 digits...." + Style.RESET_ALL)
    except Exception as e:
        print("Error found!!, ", e)
