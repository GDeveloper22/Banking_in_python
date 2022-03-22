import openpyxl.styles
from colorama import Fore, Style, init
import datetime

init(autoreset=True)
ws = openpyxl.load_workbook("./Account Database.xlsx")
wsa = ws.active

global f_name, m_name, l_name, dob, age, gender, email, pn, emp_status, ini_dep, acc_type, acc_pin, acc_status, acc_no, T_C, acc_row, ac_row


def deposit():
    # ask for account number, then check if it is numeric, then check in the database, save the row number, ask for the access pin then watch if it is numeric, if yes then ask for amount to deposit, then check if it is >1 and numeric then add the amount in the account. including the account activity, else ask to enter pin again.
    try:
        global acc_row
        while True:
            print(Style.DIM + Fore.WHITE + "Enter your account number: ", end="" + Style.RESET_ALL)
            acc_number = input()
            if len(acc_number) == 7:
                if acc_number.isnumeric():
                    for i in range(2, wsa.max_row+1):
                        if acc_number == str(wsa.cell(i, 1).value):
                            acc_row = i
                            max_attempt = 3
                            acc_stat = wsa.cell(acc_row, 3).value
                            if acc_stat == "Active":
                                while max_attempt > 0:
                                    print(Style.DIM + Fore.WHITE + "Enter the access pin: ", end="" + Style.RESET_ALL)
                                    acc_pin_dep = input()
                                    if acc_pin_dep.isnumeric():
                                        if len(acc_pin_dep) == 4:
                                            if acc_pin_dep == str(wsa.cell(acc_row, 2).value):
                                                while True:
                                                    print(Style.DIM + Fore.WHITE + "Enter the amount you want to deposit: ", end="" + Style.RESET_ALL)
                                                    acc_dep = input()
                                                    if acc_dep.isnumeric():
                                                        while True:
                                                            print(Style.DIM + Fore.WHITE + "Confirm your deposit... (Y/N): ", end="" + Style.RESET_ALL)
                                                            confirm = input()
                                                            if confirm.upper() == "Y":
                                                                wsa.cell(acc_row, 14).value = int(wsa.cell(acc_row, 14).value) + int(acc_dep)
                                                                wsa.cell(acc_row, 16).value = str(wsa.cell(acc_row, 16).value + str(f"\n${acc_dep} has been deposited in to your account on {datetime.datetime.now()}"))
                                                                ws.save("Account Database.xlsx")
                                                                print(Style.DIM + Fore.RED + f"Amount deposited in the account {acc_number}" + Style.RESET_ALL)
                                                                return 0
                                                            elif confirm.upper() == "N":
                                                                print(Style.DIM + Fore.RED + "Deposit Declined..." + Style.RESET_ALL)
                                                                return -1
                                                            else:
                                                                print(Style.DIM + Fore.RED + "Enter valid input" + Style.RESET_ALL)
                                                    else:
                                                        print(Style.DIM + Fore.RED + "Enter valid deposit amount..." + Style.RESET_ALL)
                                            else:
                                                max_attempt -= 1
                                                if max_attempt == 0:
                                                    print(Style.DIM + Fore.RED + f"({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                                else:
                                                    print(Style.DIM + Fore.RED + f"Please enter correct access pin ({max_attempt} Attempts left)..." + Style.RESET_ALL)
                                        else:
                                            print(Style.DIM + Fore.RED + "The pin should be 4 digit only..." + Style.RESET_ALL)
                                    else:
                                        print(Style.DIM + Fore.RED + "The pin should contains only digit..." + Style.RESET_ALL)
                                print(Style.DIM + Fore.RED + "Out of attempts..." + Style.RESET_ALL)
                                return 0
                            else:
                                print(Style.DIM + Fore.RED + "This account is Inactive, due to which you can not deposit..." + Style.RESET_ALL)
                                return 0
                    print(Style.DIM + Fore.RED + "Account number not found..." + Style.RESET_ALL)
                else:
                    print(Style.DIM + Fore.RED + "Account number must contain digits only...." + Style.RESET_ALL)
            else:
                print(Style.DIM + Fore.RED + "Account number must be of 7 digits...." + Style.RESET_ALL)
    except Exception as e:
        print("Error found!!, ", e)
